"""Load and normalize the employee roster from XLSX into a tree structure."""

from pathlib import Path
import openpyxl


def load_roster(xlsx_path: Path) -> dict:
    """
    Read roster XLSX and return the org tree as a nested dict.

    The root node is the person whose manager is not listed in the NAME column.
    If that person is not listed as a NAME themselves (e.g. the VP above the
    team), a synthetic root node is created for them.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Detect column positions from header row (supports both 2-col and 3-col exports)
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]
    try:
        name_col     = headers.index("NAME")
        reports_col  = headers.index("REPORTS TO")
    except ValueError as e:
        raise ValueError(f"roster.xlsx is missing expected column: {e}") from e

    # Pass 1: collect all named employees
    employees = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name      = row[name_col]
        reports_to = row[reports_col]
        # Skip empty rows and summary/footer rows (e.g. "Count Of Employees In Report:38")
        if not name or str(name).startswith("Count Of"):
            continue
        employees[name] = {
            "name": name,
            "reports_to": reports_to,
            "children": [],
        }

    # Pass 2: collect managers who appear in REPORTS TO but not in NAME
    all_managers = {emp["reports_to"] for emp in employees.values() if emp["reports_to"]}
    implicit_roots = all_managers - set(employees.keys())

    for mgr in implicit_roots:
        employees[mgr] = {
            "name": mgr,
            "reports_to": None,
            "children": [],
        }

    # Pass 3: wire up parent → children
    for emp in employees.values():
        parent = emp["reports_to"]
        if parent and parent in employees:
            employees[parent]["children"].append(emp)

    # Find the root(s)
    roots = [e for e in employees.values() if not e["reports_to"]]

    if len(roots) == 1:
        return roots[0]

    # Multiple disconnected roots → wrap in a synthetic org node
    return {
        "name": "(Organization)",
        "reports_to": None,
        "children": roots,
    }


def flatten(tree: dict) -> list[dict]:
    """Return a flat list of all employee dicts in the tree."""
    result = [tree]
    for child in tree.get("children", []):
        result.extend(flatten(child))
    return result
